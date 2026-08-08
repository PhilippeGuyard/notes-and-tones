"""Build the data files for essays/adhd/ from NHS England and NHSBSA sources.

Downloads the ADHD Management Information CSV (NHS England, May 2026 release)
and the NHSBSA Medicines Used in Mental Health BNF 4.4 workbook, caches both
under data/raw/, and emits slim per-chart JSONs into essays/adhd/data/.

Run: uv run scripts/fetch_data.py            (from tools/adhd/)
Re-run any time; outputs are deterministic for a given raw snapshot.
"""

from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests

HERE = Path(__file__).resolve().parent
RAW = HERE.parent / "data" / "raw"
OUT = HERE.parents[2] / "essays" / "adhd" / "data"

MI_URL = "https://files.digital.nhs.uk/BF/6057EB/ADHD_May26_V2.csv"
MUMH_URL = "https://nhsbsa-opendata.s3.eu-west-2.amazonaws.com/mumh/mumh_bnf0404_2022_23_v001.xlsx"

MI_SOURCE = ("NHS England, ADHD Management Information, May 2026 release "
             "(digital.nhs.uk/data-and-information/publications/statistical/mi-adhd)")
MUMH_SOURCE = ("NHSBSA, Medicines Used in Mental Health 2015/16 to 2022/23, "
               "BNF 4.4 CNS stimulants and drugs used for ADHD, identified patients")

REF_MONTH = date(2026, 3, 1)          # latest complete month in the May 2026 release
PREV_MONTH = date(2025, 3, 1)         # a year earlier, for new-referral growth
FY_FROM, FY_TO = "2019/2020", "2022/2023"


def fetch(url: str, name: str) -> Path:
    p = RAW / name
    if not p.exists():
        RAW.mkdir(parents=True, exist_ok=True)
        print(f"downloading {url} ...")
        r = requests.get(url, timeout=120)
        r.raise_for_status()
        p.write_bytes(r.content)
    return p


def parse_period(s: str) -> date | None:
    # MHSDS rows use ISO dates, CHS SitRep rows use dd/mm/yyyy.
    for f in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def mi_values(path: Path) -> dict:
    """Sum national values per (period, indicator) over the Age Group breakdown,
    plus the Gender split of open referrals for the reference month."""
    by_period: dict[tuple[date, str], int] = {}
    gender: dict[str, int] = {}
    with path.open(newline="") as f:
        for row in csv.DictReader(f):
            period = parse_period(row["REPORTING_PERIOD_START_DATE"])
            if period is None:
                continue
            ind = row["INDICATOR_ID"]
            if not row["VALUE"].isdigit():   # '*' marks suppressed small counts
                continue
            val = int(row["VALUE"])
            if row["BREAKDOWN"] == "Age Group":
                key = (period, ind)
                by_period[key] = by_period.get(key, 0) + val
            elif (row["BREAKDOWN"] == "Gender" and ind == "ADHD003"
                  and period == REF_MONTH):
                gender[row["PRIMARY_LEVEL"]] = val
    return {"by_period": by_period, "gender": gender}


def build_funnel(mi: dict) -> dict:
    v = lambda ind, period=REF_MONTH: mi["by_period"].get((period, ind), 0)
    # The prevalence estimate is stamped with the publication month, not REF_MONTH.
    prev_period = max(p for (p, ind) in mi["by_period"] if ind == "ADHD001")
    prevalence = v("ADHD001", prev_period)
    mhsds, chs = v("ADHD003"), v("ADHD008")
    new_now, new_prev = v("ADHD007"), v("ADHD007", PREV_MONTH)
    return {
        "source": MI_SOURCE,
        "retrieved": str(date.today()),
        "note": ("England, March 2026. Referral counts are 'up to' figures: MHSDS "
                 "referrals may include non-ADHD assessments, and the CHS community "
                 "paediatrics figure is an estimate. Independent providers were added "
                 "to the collection in February 2025, so comparisons across that "
                 "point partly reflect the change in coverage. CHS waiting bands "
                 "start at 12 weeks, MHSDS bands at 13."),
        "data": {
            "asof": str(REF_MONTH),
            "estimated_prevalence": prevalence,
            "open_referrals_total": mhsds + chs,
            "mhsds": {
                "open": mhsds,
                "waited": {"lt13w": v("ADHD003a"), "w13to52": v("ADHD003b"),
                           "w52to104": v("ADHD003c"), "over104w": v("ADHD003d")},
                "no_contact_yet": v("ADHD004"),
                "had_first_contact": v("ADHD005"),
            },
            "chs_estimated": {
                "open": chs,
                "waited": {"lt12w": v("ADHD008a"), "w12to52": v("ADHD008b"),
                           "w52to104": v("ADHD008c"), "over104w": v("ADHD008d")},
            },
            "over_2y_total": v("ADHD003d") + v("ADHD008d"),
            "new_referrals_month": new_now,
            "new_referrals_yoy_pct": round(100 * (new_now - new_prev) / new_prev, 1),
            "open_gender": {"male": mi["gender"].get("1", 0),
                            "female": mi["gender"].get("2", 0)},
        },
    }


def build_prescribing(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True)

    def patients(sheet, *match):
        out = {}
        for r in sheet.iter_rows(values_only=True):
            if r[0] in (FY_FROM, FY_TO) and "Y" in r and all(m in r for m in match):
                out[r[0]] = out.get(r[0], 0) + [c for c in r if isinstance(c, (int, float))][-3]
        return out

    bands = {}
    for r in wb["Age_Band"].iter_rows(values_only=True):
        if r[0] in (FY_FROM, FY_TO) and r[4] == "Y":
            bands.setdefault(r[3], {})[r[0]] = r[5]

    def growth(a, b):
        return round(100 * (b - a) / a, 1)

    keep = ["05-09", "10-14", "15-19", "20-24", "25-29", "30-34",
            "35-39", "40-44", "45-49", "50-54", "55-59", "60-64"]
    over65 = [b for b in bands if b >= "65" and b != "Unknown"]
    band_rows = [{"band": b.replace("-", " to "),
                  "from": bands[b][FY_FROM], "to": bands[b][FY_TO],
                  "growth_pct": growth(bands[b][FY_FROM], bands[b][FY_TO])}
                 for b in keep]
    f65 = sum(bands[b][FY_FROM] for b in over65)
    t65 = sum(bands[b][FY_TO] for b in over65)
    band_rows.append({"band": "65+", "from": f65, "to": t65, "growth_pct": growth(f65, t65)})

    tot_f = sum(v[FY_FROM] for v in bands.values())
    tot_t = sum(v[FY_TO] for v in bands.values())

    g3034 = {}
    for r in wb["Age_Band_and_Gender"].iter_rows(values_only=True):
        if r[0] in (FY_FROM, FY_TO) and r[3] == "30-34" and r[5] == "Y":
            g3034.setdefault(r[4], {})[r[0]] = r[6]

    return {
        "source": MUMH_SOURCE,
        "retrieved": str(date.today()),
        "note": ("Patients identified by NHS number on at least one prescription for a "
                 "BNF 4.4 medicine dispensed in the community in England, by financial "
                 "year. Growth compares 2019/20 with 2022/23."),
        "data": {
            "period_from": "2019/20", "period_to": "2022/23",
            "overall": {"from": tot_f, "to": tot_t, "growth_pct": growth(tot_f, tot_t)},
            "bands": band_rows,
            "gender_30_34": {
                s.lower(): {"from": g3034[s][FY_FROM], "to": g3034[s][FY_TO],
                            "growth_pct": growth(g3034[s][FY_FROM], g3034[s][FY_TO])}
                for s in ("Female", "Male")
            },
        },
    }


def main() -> None:
    mi_path = fetch(MI_URL, "ADHD_May26_V2.csv")
    mumh_path = fetch(MUMH_URL, "mumh_bnf0404_2022_23.xlsx")
    OUT.mkdir(parents=True, exist_ok=True)
    for name, payload in [("funnel", build_funnel(mi_values(mi_path))),
                          ("prescribing", build_prescribing(mumh_path))]:
        p = OUT / f"{name}.json"
        p.write_text(json.dumps(payload, indent=2) + "\n")
        print(f"wrote {p}")


if __name__ == "__main__":
    main()
